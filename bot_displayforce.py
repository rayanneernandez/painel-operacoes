#!/usr/bin/env python3
"""
bot_displayforce.py
══════════════════════════════════════════════════════════════
  BOT DISPLAYFORCE → PAINEL DE OPERAÇÕES
  Sincronização de campanhas UMA VEZ por dia às 09:00
══════════════════════════════════════════════════════════════

Uso:
    python bot_displayforce.py              # agenda e roda às 09:00
    python bot_displayforce.py --agora      # executa imediatamente
    python bot_displayforce.py --agora --headless=false  # navegador visível

Dependências:
    pip install playwright supabase schedule requests openpyxl pandas
    playwright install chromium
"""

import email as email_lib
import imaplib
import io
import json
import logging
import os
import re
import sys
import threading
import time
import warnings
import zipfile
from collections import defaultdict
from datetime import datetime, timezone, timedelta, date
from pathlib import Path

import requests
import schedule

# ── Suprime erros de pipe do asyncio no Windows (harmless, mas poluem o terminal) ─
# IMPORTANTE: NÃO usar WindowsSelectorEventLoopPolicy — ela quebra subprocess no Windows
# O ProactorEventLoop (padrão no Python 3.8+ Windows) é necessário para o Playwright funcionar
if sys.platform == "win32":
    warnings.filterwarnings("ignore", message=".*I/O operation on closed.*")
    warnings.filterwarnings("ignore", message=".*Exception ignored.*")
    # Força o encoding UTF-8 no stdout/stderr para evitar erros com emojis no Windows
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# ── Carrega .env ───────────────────────────────────────────────────────────────

def _carregar_env_arquivo(caminho: str = ".env", forcar: bool = False):
    if not os.path.exists(caminho):
        return
    with open(caminho, encoding="utf-8") as f:
        for linha in f:
            linha = linha.strip()
            if not linha or linha.startswith("#"):
                continue
            if "=" not in linha:
                continue
            k, v = linha.split("=", 1)
            k = k.strip()
            v = v.strip().strip('"').strip("'")
            # Suporte a comentário inline: senha-ou-app-password
            v = v.split("#")[0].strip()
            # forcar=True garante que o .env sempre sobrescreve variáveis do sistema Windows
            if k and (forcar or k not in os.environ):
                os.environ[k] = v

_carregar_env_arquivo(".env.local", forcar=True)
_carregar_env_arquivo(".env", forcar=True)   # sempre sobrescreve env vars do Windows


def _env(chave: str, padrao: str = "") -> str:
    """Lê variável de ambiente; usa default se não existir."""
    return os.environ.get(chave, padrao).strip()


# ── Variáveis globais (podem ser sobrescritas pelo Supabase) ──────────────────

DISPLAYFORCE_LOGIN_URL = "https://id.displayforce.ai/#/login"
DISPLAYFORCE_PLATFORMS_URL = "https://id.displayforce.ai/#/platforms"

DISPLAYFORCE_EMAIL = _env("DISPLAYFORCE_EMAIL")
DISPLAYFORCE_PASS  = _env("DISPLAYFORCE_PASS")

RELATORIO_EMAIL = _env("RELATORIO_EMAIL") or _env("IMAP_EMAIL")
IMAP_EMAIL      = _env("IMAP_EMAIL")
IMAP_PASSWORD   = _env("IMAP_PASSWORD")
IMAP_SERVER     = _env("IMAP_SERVER", "imap.gmail.com")
IMAP_PORT       = int(_env("IMAP_PORT", "993"))

SUPABASE_URL = _env("SUPABASE_URL") or _env("VITE_SUPABASE_URL")
SUPABASE_KEY = (
    _env("SUPABASE_SERVICE_ROLE_KEY")
    or _env("SUPABASE_KEY")
    or _env("VITE_SUPABASE_ANON_KEY")
)

HORARIO_EXECUCAO   = _env("HORARIO_EXECUCAO", "09:00")
TIMEOUT_EMAIL_SEG  = int(_env("TIMEOUT_EMAIL_SEG", "600"))   # 10 min padrão
HEADLESS_DEFAULT   = _env("HEADLESS", "true").lower() not in ("false", "0", "no")
CLIENTES_FALLBACK  = _env("CLIENTES_FALLBACK", "")           # nome|uuid,nome2|uuid2

DOWNLOAD_DIR = Path(_env("DOWNLOAD_DIR", "./downloads_displayforce"))
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

_bot_lock = threading.Lock()

# ── Logging ────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("bot_displayforce.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("bot")


# ── Supabase REST ──────────────────────────────────────────────────────────────

def _sb_headers():
    """Headers padrão para chamadas REST ao Supabase."""
    return {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }


def _sb_get(tabela: str, params: dict = None) -> list:
    """GET na REST API do Supabase usando requests (lê proxy do Windows automaticamente).
    Retorna lista de registros ou lança exceção."""
    url = f"{SUPABASE_URL}/rest/v1/{tabela}"
    r = requests.get(url, headers=_sb_headers(), params=params, timeout=30)
    r.raise_for_status()
    return r.json()


def _sb_post(tabela: str, payload, on_conflict: str = None):
    """POST (upsert) na REST API do Supabase usando requests."""
    url = f"{SUPABASE_URL}/rest/v1/{tabela}"
    headers = {**_sb_headers(), "Prefer": "resolution=merge-duplicates,return=representation"}
    params = {}
    if on_conflict:
        params["on_conflict"] = on_conflict
    if isinstance(payload, list):
        data = json.dumps(payload)
    else:
        data = json.dumps(payload)
    r = requests.post(url, headers=headers, params=params, data=data, timeout=30)
    if not r.ok:
        raise requests.HTTPError(
            f"{r.status_code} Client Error: {r.text[:300]}",
            response=r
        )
    r.raise_for_status()
    try:
        return r.json()
    except Exception:
        return []


def _sb_patch(tabela: str, filtros: dict, payload: dict):
    """PATCH (update) na REST API do Supabase usando requests."""
    url = f"{SUPABASE_URL}/rest/v1/{tabela}"
    headers = {**_sb_headers(), "Prefer": "return=representation"}
    r = requests.patch(url, headers=headers, params=filtros, json=payload, timeout=30)
    r.raise_for_status()
    return r.json()


def _sb_delete(tabela: str, filtros: dict):
    """DELETE na REST API do Supabase."""
    url = f"{SUPABASE_URL}/rest/v1/{tabela}"
    r = requests.delete(url, headers=_sb_headers(), params=filtros, timeout=30)
    r.raise_for_status()


# ── Config do Supabase ─────────────────────────────────────────────────────────

def carregar_config_supabase():
    """Lê o agendamento da tabela `bot_configs` no Supabase e aplica nos
    valores globais. Credenciais de acesso ficam no .env.
    Usa requests (lê proxy do Windows automaticamente)."""
    global HORARIO_EXECUCAO, TIMEOUT_EMAIL_SEG

    if not SUPABASE_URL or not SUPABASE_KEY:
        log.info("  Supabase não configurado — usando agendamento do .env")
        return

    try:
        rows = _sb_get("bot_configs", {"select": "horario_execucao,timeout_email_seg", "limit": "1"})
        if not rows:
            log.info("  Nenhuma config de agendamento salva no Supabase — usando .env")
            return
        cfg = rows[0]
        if cfg.get("horario_execucao"):
            HORARIO_EXECUCAO = str(cfg["horario_execucao"])
        if cfg.get("timeout_email_seg"):
            TIMEOUT_EMAIL_SEG = int(cfg["timeout_email_seg"])
        log.info(f"  ✅ Agendamento carregado do Supabase: {HORARIO_EXECUCAO} | timeout={TIMEOUT_EMAIL_SEG}s")
    except Exception as e:
        log.warning(f"  Não foi possível carregar agendamento do Supabase: {e}")


# ── Clientes ───────────────────────────────────────────────────────────────────

def _clientes_fallback() -> list[dict]:
    """Retorna clientes a partir da variável CLIENTES_FALLBACK no .env.
    Formato: nome1|uuid1,nome2|uuid2
    Exemplo: assai|c6999bd9-14c0-4e26-abb1-d4b8xxxx,panvel|b1c05e4d-0417-4853-9af9-8c0725df18"""
    if not CLIENTES_FALLBACK:
        return []
    clientes = []
    for parte in CLIENTES_FALLBACK.split(","):
        parte = parte.strip()
        if "|" in parte:
            nome, uid = parte.split("|", 1)
            clientes.append({"id": uid.strip(), "name": nome.strip()})
    log.info(f"  📋 Clientes do fallback (.env): {[c['name'] for c in clientes]}")
    return clientes


def buscar_clientes() -> list[dict]:
    """Busca clientes ativos no Supabase via REST (usa proxy do Windows automaticamente)."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        return _clientes_fallback()
    try:
        for status in ("active", "ativo"):
            rows = _sb_get("clients", {"select": "id,name,status", "status": f"eq.{status}"})
            if rows:
                log.info(f"  Clientes encontrados no Supabase (status={status}): {[r['name'] for r in rows]}")
                return rows
        # Fallback: todos os clientes
        rows = _sb_get("clients", {"select": "id,name"})
        if rows:
            log.warning(f"  Nenhum cliente com status 'active/ativo'. Usando todos.")
            return rows
        log.error("  Nenhum cliente encontrado no Supabase.")
        return _clientes_fallback()
    except Exception as e:
        log.error(f"  Erro ao buscar clientes no Supabase: {e}")
        return _clientes_fallback()


# ── Upsert de campanhas ────────────────────────────────────────────────────────

_PENDING_DIR = Path("./pending_uploads")


def upsert_campanhas(client_id: str, client_name: str, registros: list[dict]):
    """Salva campanhas no Supabase via REST. Em caso de falha, salva localmente."""

    def _norm_text(v):
        if v is None:
            return ""
        return str(v).strip().lower()

    def _campaign_match_variants(r: dict) -> list[str]:
        """Gera variantes de chave para dedup."""
        variants = []
        nome      = _norm_text(r.get("name"))
        content   = _norm_text(r.get("content_name"))
        tipo      = _norm_text(r.get("tipo_midia"))
        loja      = _norm_text(r.get("loja"))
        variants.append(f"{nome}|{content}|{tipo}|{loja}")
        variants.append(f"{nome}|{content}|{tipo}|")
        variants.append(f"{nome}|{content}||")
        variants.append(f"{nome}|||")
        return variants

    if not registros:
        log.warning(f"  Nenhuma campanha para salvar para '{client_name}'")
        return

    try:
        # SEM flush — dados de meses anteriores são preservados.
        # UPSERT: insere campanha nova, substitui se a mesma chave já existe
        # Chave única: (client_id, name, content_name, tipo_midia, loja)
        total = 0
        erros = 0
        for i in range(0, len(registros), 200):
            chunk = registros[i:i+200]
            try:
                res = _sb_post("campaigns", chunk, on_conflict="client_id,name,content_name,tipo_midia,loja")
                total += len(res) if isinstance(res, list) else len(chunk)
            except Exception as e_upsert:
                log.warning(f"  Upsert falhou ({e_upsert}) — tentando INSERT direto...")
                try:
                    _sb_post("campaigns", chunk)
                    total += len(chunk)
                except Exception as e_insert:
                    log.error(f"  Insert lote {i//200+1} falhou: {e_insert}")
                    erros += len(chunk)
        if erros:
            log.warning(f"  ⚠️  {erros} registros não inseridos por erro")
        log.info(f"  ✅ {total} campanhas salvas/atualizadas no Supabase")

    except Exception as e:
        log.error(f"  ❌ Supabase indisponível — salvando dados localmente para reenvio posterior")
        _PENDING_DIR.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        arquivo = _PENDING_DIR / f"pendente_{client_id}_{ts}.json"
        arquivo.write_text(json.dumps({"client_id": client_id, "registros": registros}, ensure_ascii=False), encoding="utf-8")
        log.info(f"  💾 Dados salvos em: {arquivo} — serão reenviados na próxima execução")


def _reenviar_pendentes():
    """Tenta reenviar arquivos JSON pendentes para o Supabase."""
    if not _PENDING_DIR.exists():
        return
    arquivos = list(_PENDING_DIR.glob("pendente_*.json"))
    if not arquivos:
        return
    log.info(f"  📤 {len(arquivos)} arquivo(s) pendente(s) para reenvio...")
    for arq in arquivos:
        try:
            dados = json.loads(arq.read_text(encoding="utf-8"))
            _sb_delete("campaigns", {"client_id": f"eq.{dados['client_id']}"})
            _sb_post("campaigns", dados["registros"], "client_id,name,content_name,tipo_midia,loja")
            arq.unlink()
            log.info(f"  ✅ Pendente reenviado e removido: {arq.name}")
        except Exception as e:
            log.warning(f"  ⚠️  Falha ao reenviar {arq.name}: {e}")


# ── Parsing do relatório ───────────────────────────────────────────────────────

def parse_tempo_atencao(valor) -> float | None:
    """Converte string de tempo (ex: '00:01:30' ou '90') para segundos."""
    if valor is None:
        return None
    s = str(valor).strip()
    if not s or s == "-":
        return None
    # Formato HH:MM:SS
    m = re.match(r"(\d+):(\d{2}):(\d{2})", s)
    if m:
        return int(m.group(1)) * 3600 + int(m.group(2)) * 60 + int(m.group(3))
    try:
        return float(s)
    except ValueError:
        return None


def parse_data(valor) -> str | None:
    """Normaliza data para ISO string ou None."""
    if valor is None:
        return None
    s = str(valor).strip()
    if not s or s == "-":
        return None
    try:
        from datetime import datetime as dt
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
            try:
                return dt.strptime(s, fmt).isoformat()
            except ValueError:
                continue
    except Exception:
        pass
    return s


def _clean_content_name(raw: str) -> str:
    """Limpa o nome do vídeo/conteúdo removendo:
    - Períodos de exibição (ex: _10mar_15mar_26, -05mar_26-30mai_26)
    - Resolução e formato (ex: - 1080x1920, _1080x1920, (43' Vertical))
    - Extensão .mp4
    """
    name = str(raw).strip()
    name = re.sub(r"\.mp4$", "", name, flags=re.IGNORECASE)
    meses = r"(?:jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)"
    name = re.sub(rf"(?:[_\-\s]+\d{{1,2}}{meses})+(?:[_\-]\d{{2,4}})?", "", name, flags=re.IGNORECASE)
    name = re.sub(r"[_\-\s]*\d{3,4}\s*[xX]\s*\d{3,4}", "", name)
    name = re.sub(r"\s*\([^)]*(?:vertical|horizontal|vert|horiz)[^)]*\)", "", name, flags=re.IGNORECASE)
    name = re.sub(r"\s*\(\d+\)\s*$", "", name)
    name = re.sub(r"[_\s]+v\d+$", "", name, flags=re.IGNORECASE)
    return re.sub(r"[_\-\s]+$", "", name).strip()


def processar_excel(caminho: str, client_id: str) -> list[dict]:
    """Processa Excel legado (aba 'Visitors' ou 'Report') — fallback caso não haja Views CSV."""
    import openpyxl
    log.info(f"  Processando Excel: {caminho}")
    try:
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    except Exception as e:
        log.error(f"  Erro ao ler arquivo: {e}")
        return []

    # Tenta aba de Visitors para compatibilidade antiga
    aba = None
    for nome in ("Views of visitors", "Visitors", "Campanhas", wb.sheetnames[0]):
        if nome in wb.sheetnames:
            aba = wb[nome]
            break
    if aba is None:
        return []

    linhas = list(aba.iter_rows(values_only=True))
    if not linhas:
        return []

    headers = [str(c or "").strip() for c in linhas[0]]
    log.info(f"  Colunas encontradas: {headers[:20]}")

    def encontrar_coluna(*candidatos) -> int | None:
        hl = [h.lower() for h in headers]
        for c in candidatos:
            for i, h in enumerate(hl):
                if c.lower() in h:
                    return i
        return None

    col_campaign = encontrar_coluna("campaign", "campanha", "nome")
    col_content  = encontrar_coluna("content", "conteudo", "video")
    col_device   = encontrar_coluna("device", "dispositivo", "loja")
    col_visitors = encontrar_coluna("visitor", "visitante")
    col_atencao  = encontrar_coluna("attention", "atencao", "duration")
    col_start    = encontrar_coluna("start", "inicio", "date")
    col_end      = encontrar_coluna("end", "fim")

    mapeamento = {
        "campaign": col_campaign, "content": col_content,
        "device": col_device, "visitors": col_visitors,
        "atencao": col_atencao, "start": col_start, "end": col_end,
    }
    log.info(f"  Mapeamento de colunas: {mapeamento}")

    agr: dict = defaultdict(lambda: {"visitors": set(), "display_count": 0, "attn_sum": 0.0, "attn_count": 0, "start": None, "end": None})
    agora = datetime.now(timezone.utc).isoformat()

    for row in linhas[1:]:
        def _get(idx): return str(row[idx]).strip() if idx is not None and idx < len(row) and row[idx] is not None else ""
        campaign = _get(col_campaign)
        content  = _clean_content_name(_get(col_content)) if col_content is not None else None
        device   = _get(col_device)
        visitor  = _get(col_visitors)

        if not campaign and not content:
            continue

        loja, tipo_midia = _parse_device(device) if device else ("", "")
        key = (campaign or content, content or "", tipo_midia, loja)
        rec = agr[key]

        if visitor:
            rec["visitors"].add(visitor)
        rec["display_count"] += 1

        attn = parse_tempo_atencao(_get(col_atencao) if col_atencao is not None else None)
        if attn and attn > 0:
            rec["attn_sum"] += attn
            rec["attn_count"] += 1

        s = parse_data(_get(col_start) if col_start is not None else None)
        e = parse_data(_get(col_end) if col_end is not None else None)
        if s and (rec["start"] is None or s < rec["start"]): rec["start"] = s
        if e and (rec["end"] is None or e > rec["end"]): rec["end"] = e

    registros = []
    for (name, content_name, tipo_midia, loja), rec in agr.items():
        registros.append({
            "client_id": client_id,
            "name": name,
            "content_name": content_name or "",   # NOT NULL na DB
            "tipo_midia": tipo_midia or "",        # NOT NULL na DB
            "loja": loja or "",                    # NOT NULL na DB
            "start_date": rec["start"],
            "end_date": rec["end"],
            "duration_days": None,
            "duration_hms": None,
            "display_count": rec["display_count"],
            "visitors": len(rec["visitors"]),
            "avg_attention_sec": int(rec["attn_sum"] / rec["attn_count"]) if rec["attn_count"] else 0,
            "uploaded_at": agora,
        })
    log.info(f"  {len(registros)} linhas extraídas do Excel")
    return registros


def _parse_device(device_name: str) -> tuple[str, str]:
    """Retorna (loja, tipo_midia) a partir do nome do Device."""
    if " - " in device_name:
        loja, tipo = device_name.rsplit(" - ", 1)
        return loja.strip(), tipo.strip()
    return device_name.strip(), ""


def processar_views_csv(caminho: str, client_id: str) -> list[dict]:
    """Lê o arquivo 'Views of visitors' CSV do DisplayForce e agrega por
    Campaign + Device → salva como (name, tipo_midia, loja) na tabela campaigns.

    Estrutura confirmada do CSV DisplayForce:
      Campaign | Content name | Device | Duration (seconds) | ...
    """
    import csv
    log.info(f"  Processando Views CSV: {caminho}")
    try:
        with open(caminho, encoding="utf-8-sig", errors="replace") as f:
            # Pula linha "sep=," que o Excel exporta antes do cabeçalho real
            primeira = f.readline()
            if primeira.strip().strip('"').lower() != "sep=,":
                f.seek(0)  # não era sep=, — volta ao início
            reader = csv.DictReader(f)
            linhas = list(reader)
    except Exception as e:
        log.error(f"  Erro ao ler Views CSV: {e}")
        return []

    if not linhas:
        return []

    headers = list(linhas[0].keys())
    log.info(f"  Colunas Views CSV ({len(headers)}): {headers[:15]}")
    log.info(f"  Total de linhas brutas: {len(linhas)}")

    def _find_col(*candidatos) -> str | None:
        hl = {h.lower().strip(): h for h in headers if h is not None}
        # Prioridade 1: match exato
        for c in candidatos:
            if c.lower().strip() in hl:
                return hl[c.lower().strip()]
        # Prioridade 2: match parcial (candidato contido no header)
        for c in candidatos:
            c_lower = c.lower().strip()
            for h_lower, h_orig in hl.items():
                if c_lower in h_lower:
                    return h_orig
        return None

    # Nomes exatos primeiro para evitar "Campaign ID" no lugar de "Campaign"
    col_campaign = _find_col("campaign", "campanha")
    col_content  = _find_col("content", "content name", "conteudo", "video")
    col_device   = _find_col("device", "dispositivo")
    col_duration = _find_col("content view duration", "duration", "atencao", "attention", "dwell")
    col_visitor  = _find_col("visitor id", "visitor", "visitante")
    col_start    = _find_col("content view start", "start", "inicio", "played_at", "timestamp")
    col_end      = _find_col("content view end", "end", "fim")
    col_gender   = _find_col("gender", "genero", "sexo")
    col_age      = _find_col("age", "idade")

    log.info(f"  Mapeamento: campaign={col_campaign}, content={col_content}, device={col_device}, duration={col_duration}, gender={col_gender}, age={col_age}")

    if not col_campaign and not col_content:
        log.error("  Colunas 'Campaign' ou 'Content' não encontradas — abortando Views CSV")
        return []

    def _age_label(age_val: str) -> str | None:
        try:
            a = int(float(age_val))
            if a <= 0: return None
            if a < 18:  return "0-17"
            if a < 25:  return "18-24"
            if a < 35:  return "25-34"
            if a < 45:  return "35-44"
            if a < 55:  return "45-54"
            return "55+"
        except (ValueError, TypeError):
            return None

    agr: dict = defaultdict(lambda: {
        "visitors": set(), "display_count": 0,
        "attn_sum": 0.0, "attn_count": 0,
        "start": None, "end": None,
        "gender": {"male": 0, "female": 0, "unknown": 0},
        "age": {},
    })
    agora = datetime.now(timezone.utc).isoformat()

    linhas_uteis = 0
    for row in linhas:
        campaign   = str(row.get(col_campaign) or "").strip() if col_campaign else ""
        content    = str(row.get(col_content)  or "").strip() if col_content  else ""
        device     = str(row.get(col_device)   or "").strip() if col_device   else ""
        visitor    = str(row.get(col_visitor)  or "").strip() if col_visitor  else ""
        dur_raw    = str(row.get(col_duration) or "").strip() if col_duration else ""
        start_s    = str(row.get(col_start)    or "").strip() if col_start    else ""
        end_s      = str(row.get(col_end)      or "").strip() if col_end      else ""
        gender_raw = str(row.get(col_gender)   or "").strip().lower() if col_gender else ""
        age_raw    = str(row.get(col_age)      or "").strip() if col_age      else ""

        if not campaign and not content:
            continue

        content_clean = _clean_content_name(content) if content else None
        loja, tipo_midia = _parse_device(device) if device else ("", "")
        name = campaign or content_clean or content

        key = (name, content_clean or "", tipo_midia, loja)
        rec = agr[key]

        if visitor:
            rec["visitors"].add(visitor)
        rec["display_count"] += 1

        dur = parse_tempo_atencao(dur_raw)
        if dur and dur > 0:
            rec["attn_sum"] += dur
            rec["attn_count"] += 1

        # Gênero → {"male": N, "female": N, "unknown": N}
        if "male" in gender_raw or gender_raw == "m":
            rec["gender"]["male"] += 1
        elif "female" in gender_raw or "fem" in gender_raw or gender_raw == "f":
            rec["gender"]["female"] += 1
        else:
            rec["gender"]["unknown"] += 1

        # Idade → {"18-24": N, "25-34": N, ...}
        lbl = _age_label(age_raw)
        if lbl:
            rec["age"][lbl] = rec["age"].get(lbl, 0) + 1

        s = parse_data(start_s) if start_s else None
        e = parse_data(end_s)   if end_s   else None
        if s and (rec["start"] is None or s < rec["start"]): rec["start"] = s
        if e and (rec["end"] is None or e > rec["end"]): rec["end"] = e

        linhas_uteis += 1

    if linhas_uteis == 0:
        log.warning("  Nenhuma linha útil no Views CSV")
        return []

    registros = []
    for (name, content_name, tipo_midia, loja), rec in agr.items():
        registros.append({
            "client_id": client_id,
            "name": name,
            "content_name": content_name or "",   # NOT NULL na DB
            "tipo_midia": tipo_midia or "",        # NOT NULL na DB
            "loja": loja or "",                    # NOT NULL na DB
            "start_date": rec["start"],
            "end_date": rec["end"],
            "duration_days": None,
            "duration_hms": None,
            "display_count": rec["display_count"],
            "total_play_seconds": int(rec["attn_sum"]),
            "visitors": len(rec["visitors"]),
            "avg_attention_sec": int(rec["attn_sum"] / rec["attn_count"]) if rec["attn_count"] else 0,
            "gender_breakdown": rec["gender"],
            "age_breakdown": rec["age"] if rec["age"] else None,
            "uploaded_at": agora,
        })
    log.info(f"  {len(registros)} registros extraídos do Views CSV")
    return registros


# ── IMAP ───────────────────────────────────────────────────────────────────────

def _obter_ultimo_uid(imap: imaplib.IMAP4_SSL, pasta: str = "INBOX") -> int:
    """Retorna UID real da mensagem mais recente para ignorar e-mails antigos."""
    try:
        imap.select(pasta, readonly=True)
        status, dados = imap.uid("search", None, "ALL")
        if status == "OK" and dados[0]:
            uids = dados[0].split()
            total = len(uids)
            ultimo = int(uids[-1]) if uids else 0
            log.info(f"  📬 IMAP conectado — {total} e-mails na caixa, último UID real: {ultimo}")
            return ultimo
    except Exception as e:
        log.warning(f"  Erro ao obter último UID: {e}")
    return 0


def _discover_mailboxes(imap: imaplib.IMAP4_SSL) -> list[str]:
    """Lista pastas disponíveis no servidor IMAP."""
    try:
        status, dados = imap.list()
        if status == "OK":
            pastas = []
            for d in dados:
                if d:
                    partes = d.decode("utf-8", errors="replace").split('"')
                    if partes:
                        pastas.append(partes[-1].strip().strip('"').strip())
            return [p for p in pastas if p]
    except Exception:
        pass
    return ["INBOX"]


def _extrair_link_download(msg) -> str | None:
    """Varre o corpo HTML/texto do e-mail em busca do link de download.
    Prioriza links com extensão de arquivo ou palavras-chave de relatório."""
    corpo_html = ""
    corpo_txt  = ""
    if msg.is_multipart():
        for part in msg.walk():
            ct = part.get_content_type()
            if ct == "text/html":
                corpo_html += part.get_payload(decode=True).decode("utf-8", errors="replace")
            elif ct == "text/plain":
                corpo_txt += part.get_payload(decode=True).decode("utf-8", errors="replace")
    else:
        corpo_html = msg.get_payload(decode=True).decode("utf-8", errors="replace")

    links = re.findall(r'href=["\']([^"\']+)["\']', corpo_html)
    links += re.findall(r'https?://\S+', corpo_txt)

    log.info(f"  🔗 Links candidatos encontrados: {len(links)}")

    # Prioridade 1: links com extensão de arquivo
    for link in links:
        if re.search(r"\.(xlsx|xls|csv|zip)(\?|$)", link, re.IGNORECASE):
            return link

    # Prioridade 2: links com palavras-chave de relatório
    palavras = ("report", "relatorio", "relat", "download", "export", "visitors")
    for link in links:
        if any(p in link.lower() for p in palavras):
            return link

    return None


def _extrair_zip(conteudo: bytes, destino_dir: Path) -> list[str]:
    """Extrai todos os Excel/CSV de um ZIP em memória. Retorna lista de caminhos."""
    caminhos = []
    try:
        with zipfile.ZipFile(io.BytesIO(conteudo)) as zf:
            nomes = [n for n in zf.namelist() if re.search(r"\.(xlsx|xls|csv)$", n, re.IGNORECASE)]
            if not nomes:
                log.warning(f"  ZIP não contém Excel/CSV. Arquivos dentro: {zf.namelist()}")
                return []
            for nome in nomes:
                dados = zf.read(nome)
                local = destino_dir / Path(nome).name
                local.write_bytes(dados)
                caminhos.append(str(local))
                log.info(f"  📂 Extraído do ZIP: {local.name}")
    except Exception as e:
        log.warning(f"  Erro ao extrair ZIP: {e}")
    return caminhos


def _baixar_arquivo_url(url: str, destino_dir: Path) -> list[str]:
    """Faz GET na URL e salva o(s) arquivo(s) em DOWNLOAD_DIR.
    Se for ZIP, extrai todos os Excel/CSV encontrados.
    Retorna lista de caminhos."""
    log.info(f"  ⬇️  Baixando arquivo de: {url[:100]}")
    try:
        r = requests.get(url, timeout=60, stream=True)
        r.raise_for_status()
        conteudo = r.content

        # Tenta detectar extensão pelo Content-Disposition
        cd = r.headers.get("Content-Disposition", "")
        m = re.search(r'filename=["\']?([^"\';\s]+)', cd)
        nome_arq = m.group(1) if m else "relatorio"
        ext = Path(nome_arq).suffix.lower()
        if not ext:
            ct = r.headers.get("Content-Type", "")
            if "zip" in ct: ext = ".zip"
            elif "excel" in ct or "spreadsheet" in ct: ext = ".xlsx"
            elif "csv" in ct: ext = ".csv"
            else: ext = ".xlsx"

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        log.info(f"  {len(conteudo)} bytes baixados (ext detectada: {ext})")

        if ext == ".zip":
            caminhos = _extrair_zip(conteudo, destino_dir)
            if not caminhos:
                salvo = destino_dir / f"relatorio_{ts}.zip"
                salvo.write_bytes(conteudo)
                log.warning(f"  ⚠️  ZIP sem Excel/CSV — salvo como: {salvo}")
            return caminhos

        salvo = destino_dir / f"relatorio_{ts}{ext}"
        salvo.write_bytes(conteudo)
        return [str(salvo)]

    except Exception as e:
        log.error(f"  ❌ Erro ao baixar arquivo: {e}")
        return []


def _processar_msg_displayforce(imap: imaplib.IMAP4_SSL, uid: bytes) -> list[str]:
    """Tenta extrair arquivo(s) de um e-mail e retorna lista de caminhos."""
    try:
        status, dados = imap.uid("fetch", uid, "(RFC822)")
        if status != "OK" or not dados or dados[0] is None:
            log.warning(f"  Erro ao buscar e-mail UID {uid.decode()}: fetch retornou vazio")
            return []
        if not isinstance(dados[0][1], bytes):
            log.warning(f"  Erro ao buscar e-mail UID {uid.decode()}: dados não são bytes — tipo={type(dados[0][1])}")
            return []
    except Exception as e:
        log.warning(f"  Erro ao buscar e-mail UID {uid.decode()}: {e}")
        return []

    msg = email_lib.message_from_bytes(dados[0][1])
    caminhos = []

    # Tenta anexos primeiro
    for part in msg.walk():
        nome_arq = part.get_filename()
        if not nome_arq:
            continue
        if not re.search(r"\.(xlsx|xls|csv|zip)$", nome_arq, re.IGNORECASE):
            continue
        log.info(f"  📎 Anexo encontrado: '{nome_arq}'")
        conteudo = part.get_payload(decode=True)
        if not conteudo:
            continue
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_safe = re.sub(r"[^a-zA-Z0-9._\-]", "_", nome_arq)
        salvo = DOWNLOAD_DIR / f"{ts}_{nome_safe}"
        salvo.write_bytes(conteudo)

        if nome_arq.lower().endswith(".zip"):
            extraidos = _extrair_zip(conteudo, DOWNLOAD_DIR)
            caminhos.extend(extraidos if extraidos else [str(salvo)])
        else:
            caminhos.append(str(salvo))

    if caminhos:
        return caminhos

    # Fallback: link de download no corpo
    log.info("  📧 Sem anexo — buscando link de download no corpo do e-mail...")
    link = _extrair_link_download(msg)
    if link:
        return _baixar_arquivo_url(link, DOWNLOAD_DIR)

    log.warning("  ⚠️  E-mail sem link/anexo reconhecível")
    return []


def _buscar_em_pasta(imap: imaplib.IMAP4_SSL, pasta: str, uid_inicio: int, nome_cliente: str = "") -> list[str]:
    """Tenta encontrar e processar e-mail DisplayForce numa pasta específica.
    CORRIGIDO: usa uid('search') e uid('fetch') para UIDs reais.
    nome_cliente: se informado, filtra apenas e-mails cujo remetente contenha o nome."""
    try:
        imap.select(pasta, readonly=True)
        status, dados = imap.uid("search", None, f"UID {uid_inicio}:*")
        if status != "OK" or not dados[0]:
            return []

        uids_novos = [u for u in dados[0].split() if int(u) > uid_inicio]
        log.info(f"  {len(uids_novos)} e-mail(s) na pasta '{pasta}' (UID > {uid_inicio})")
        if not uids_novos:
            return []

        for uid in reversed(uids_novos):
            try:
                status2, header_data = imap.uid("fetch", uid, "(BODY.PEEK[HEADER.FIELDS (FROM SUBJECT)])")
                if status2 != "OK" or not header_data or not header_data[0]:
                    continue
                header_bytes = header_data[0][1] if isinstance(header_data[0], tuple) else b""
                header_str = header_bytes.decode("utf-8", errors="replace").lower() if header_bytes else ""

                remetente_ok = "displ" in header_str or "displayforce" in header_str
                cliente_ok   = not nome_cliente or nome_cliente.lower() in header_str

                if not remetente_ok:
                    log.debug(f"  UID {uid.decode()} ignorado — não é DisplayForce")
                    continue
                if not cliente_ok:
                    log.debug(f"  UID {uid.decode()} — '{nome_cliente}' não está em remetente")
                    continue

                log.info(f"  ✉️  DisplayForce detectado em '{pasta}' UID {uid.decode()}")
                if pasta.upper() == "SPAM" or "[SPAM]" in pasta.upper():
                    log.warning("  ⚠️  E-mail encontrado na pasta SPAM! Mova para INBOX para evitar isso.")

                caminhos = _processar_msg_displayforce(imap, uid)
                if caminhos:
                    return caminhos

            except Exception as e:
                log.warning(f"  Erro ao processar UID {uid.decode()}: {e}")

    except Exception as e:
        log.warning(f"  Erro na pasta '{pasta}': {e}")
    return []


def _verificar_email_displayforce(uid_inicio: int, nome_cliente: str = "") -> list[str]:
    """Verifica e-mails novos da DisplayForce em INBOX e Spam.
    CORRIGIDO: usa uid('search') e uid('fetch') para UIDs reais do Gmail.
    nome_cliente: se informado, filtra apenas e-mails cujo remetente contenha o nome."""
    try:
        imap = imaplib.IMAP4_SSL(IMAP_SERVER, IMAP_PORT)
        imap.login(IMAP_EMAIL, IMAP_PASSWORD)
    except Exception as e:
        log.error(f"  ❌ Falha ao conectar ao IMAP: {e}")
        return []

    try:
        # ── Estratégia 1: busca UIDs novos em INBOX e Spam ──────────────────
        for pasta in ("INBOX", "[Gmail]/Spam", "Spam", "[Gmail]/All Mail"):
            try:
                caminhos = _buscar_em_pasta(imap, pasta, uid_inicio, nome_cliente)
                if caminhos:
                    return caminhos
            except Exception:
                continue

        # ── Estratégia 2: busca por remetente 'displ' em todo o dia ─────────
        hoje = datetime.now().strftime("%d-%b-%Y")
        log.info(f"  🔍 Sem UIDs novos — buscando por remetente 'displ' desde {hoje}")
        for pasta in ("INBOX", "[Gmail]/All Mail"):
            try:
                imap.select(pasta, readonly=True)
                status, dados = imap.uid("search", None, f'FROM "displ" SINCE {hoje}')
                if status != "OK" or not dados[0]:
                    continue
                uids = dados[0].split()
                log.info(f"  {len(uids)} e-mail(s) displ.com encontrado(s) hoje na INBOX")
                for uid in reversed(uids):
                    if int(uid) <= uid_inicio:
                        log.debug(f"  (est.2) ignorado — UID {uid.decode()} <= {uid_inicio}")
                        continue
                    caminhos = _processar_msg_displayforce(imap, uid)
                    if caminhos:
                        return caminhos
            except Exception:
                continue

        # ── Estratégia 3: busca ampla por qualquer e-mail de hoje ───────────
        log.info("  🔍 Busca ampla — qualquer e-mail de hoje com possível relatório...")
        for pasta in ("INBOX",):
            try:
                imap.select(pasta, readonly=True)
                status, dados = imap.uid("search", None, f"SINCE {hoje}")
                if status != "OK" or not dados[0]:
                    continue
                uids_hoje = [u for u in dados[0].split() if int(u) > uid_inicio]
                log.info(f"  {len(uids_hoje)} e-mail(s) de hoje não processados — verificando...")
                for uid in reversed(uids_hoje):
                    try:
                        status2, hd = imap.uid("fetch", uid, "(BODY.PEEK[HEADER.FIELDS (FROM SUBJECT)])")
                        if status2 != "OK" or not hd or not hd[0]:
                            continue
                        hb = hd[0][1] if isinstance(hd[0], tuple) else b""
                        hs = hb.decode("utf-8", errors="replace").lower()
                        # Filtra por remetente ou assunto com relatório
                        tem_relatorio = any(p in hs for p in ("displ", "visitors insight", "report", "relatorio"))
                        if not tem_relatorio:
                            log.debug(f"  (est.3) ignorado — UID {uid.decode()}")
                            continue
                        log.info(f"  🔎 Estratégia 3: UID {uid.decode()} tem cara de relatório")
                        caminhos = _processar_msg_displayforce(imap, uid)
                        if caminhos:
                            return caminhos
                    except Exception as e:
                        log.warning(f"  Erro na estratégia 3 UID {uid.decode()}: {e}")
            except Exception:
                continue

    finally:
        try:
            imap.logout()
        except Exception:
            pass

    return []


def baixar_relatorio_email(uid_inicio: int, nome_cliente: str = "",
                            timeout_seg: int = None) -> list[str]:
    """Aguarda o e-mail da DisplayForce e retorna lista de caminhos de arquivos baixados.
    Verifica a caixa IMAP a cada 15 segundos até timeout_seg.
    nome_cliente: se informado, filtra apenas e-mails cujo remetente contenha o nome."""
    if timeout_seg is None:
        timeout_seg = TIMEOUT_EMAIL_SEG

    log.info(f"  Aguardando e-mail com relatório em '{IMAP_EMAIL}' (timeout={timeout_seg}s)...")
    prazo = time.time() + timeout_seg

    while time.time() < prazo:
        caminhos = _verificar_email_displayforce(uid_inicio, nome_cliente)
        if caminhos:
            log.info(f"  ✅ {len(caminhos)} arquivo(s) baixado(s): {[Path(p).name for p in caminhos]}")
            return caminhos
        restante = int(prazo - time.time())
        if restante > 0:
            time.sleep(min(15, restante))

    log.error(f"  ⏱️  Timeout: nenhum e-mail da DisplayForce recebido em {timeout_seg}s")
    return []


# ── Playwright ─────────────────────────────────────────────────────────────────

def _screenshot(page, nome: str):
    """Salva screenshot de debug."""
    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        caminho = DOWNLOAD_DIR / f"debug_{nome}_{ts}.png"
        page.screenshot(path=str(caminho))
        log.info(f"  📸 Screenshot salvo: {caminho.name}")
    except Exception:
        pass


def _fill_input_react(page, seletor: str, valor: str, label: str = "input"):
    """Preenche input controlado por React disparando input/change com retries."""
    def _log_fill_ok(v):
        log.info(f"  ✔ {label} preenchido: '{v[:40]}'")

    for tentativa in range(3):
        try:
            el = page.locator(seletor).first
            el.wait_for(state="visible", timeout=5000)
            el.triple_click()
            el.fill(valor)
            # Dispara eventos React
            page.evaluate(
                """([el, val]) => {
                    const setter = Object.getOwnPropertyDescriptor(
                      window.HTMLInputElement.prototype,
                      'value'
                    ).set;
                    setter.call(el, val);
                    el.dispatchEvent(new Event('input', { bubbles: true }));
                    el.dispatchEvent(new Event('change', { bubbles: true }));
                }""",
                [el.element_handle(), valor]
            )
            _log_fill_ok(valor)
            return True
        except Exception as e:
            if tentativa == 2:
                log.warning(f"  Falha ao preencher {label}: {e}")
    return False


def _injetar_valor(page, seletor_css: str, valor: str):
    """Injeta valor num input via JS nativo — funciona em React, hidden e ABNT2."""
    page.evaluate(
        """([sel, val]) => {
            const el = document.querySelector(sel);
            if (!el) return;
            el.focus();
            const setter = Object.getOwnPropertyDescriptor(
                window.HTMLInputElement.prototype, 'value'
            ).set;
            setter.call(el, val);
            el.dispatchEvent(new Event('input',  { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
        }""",
        [seletor_css, valor]
    )


def _aguardar_sem_networkidle(page, ms: int = 2000):
    """Substituto seguro de wait_for_load_state('networkidle') para SPAs hash-routing."""
    try:
        page.wait_for_load_state("domcontentloaded", timeout=ms)
    except Exception:
        pass
    time.sleep(ms / 1000)


def fazer_login_displayforce(page) -> bool:
    """Realiza login na DisplayForce. Retorna True se bem-sucedido."""
    log.info("  Acessando DisplayForce...")
    try:
        page.goto(DISPLAYFORCE_LOGIN_URL, timeout=30_000)
        _aguardar_sem_networkidle(page, 2500)

        # ── Etapa 1: preenche e-mail ──────────────────────────────────────
        log.info(f"  Preenchendo e-mail: {DISPLAYFORCE_EMAIL}")
        page.wait_for_selector("input", state="visible", timeout=15_000)
        time.sleep(0.5)

        _injetar_valor(page, "input", DISPLAYFORCE_EMAIL)
        time.sleep(0.5)

        val_digitado = page.locator("input").first.input_value()
        log.info(f"  ✔ e-mail preenchido: '{val_digitado}'")

        # Clica em CONTINUAR
        page.locator(
            "button:has-text('CONTINUAR'), button:has-text('Continuar'), button:has-text('CONTINUE')"
        ).first.click()
        log.info("  Botão CONTINUAR clicado")

        # Espera SPA transicionar — SEM networkidle (hash routing trava o networkidle no Windows)
        time.sleep(3)
        _screenshot(page, "login_02_apos_proximo")

        # ── Etapa 2: preenche senha ───────────────────────────────────────
        log.info("  Aguardando campo de senha...")
        # Aguarda o campo existir no DOM — pode estar hidden (classe pseudoPassword__bas2)
        try:
            page.wait_for_selector("input[type='password']", state="attached", timeout=12_000)
        except Exception:
            time.sleep(3)   # Dá mais tempo e continua mesmo assim

        time.sleep(0.5)
        # Injeta senha via JS — funciona mesmo que o campo esteja hidden
        _injetar_valor(page, 'input[type="password"]', DISPLAYFORCE_PASS)
        time.sleep(0.5)
        log.info(f"  ✔ senha preenchida (len={len(DISPLAYFORCE_PASS)})")

        # Clica em ENTRAR / submit
        for seletor in [
            "button:has-text('ENTRAR')", "button:has-text('Entrar')",
            "button:has-text('CONTINUAR')", "button:has-text('Continuar')",
            "button:has-text('LOGIN')", "button[type='submit']",
        ]:
            try:
                btn = page.locator(seletor).first
                if btn.count() > 0:
                    btn.click(timeout=3000)
                    log.info(f"  ✔ Botão login clicado: {seletor}")
                    break
            except Exception:
                continue

        # Aguarda navegação — SEM networkidle
        time.sleep(4)
        _screenshot(page, "login_03_pos_login")

        # Verifica sucesso
        url_atual = page.url
        log.info(f"  URL após login: {url_atual}")
        if "platform" in url_atual or "login" not in url_atual:
            log.info("  ✅ Login realizado com sucesso")
            return True
        else:
            log.error(f"  ❌ Login falhou — ainda na tela: {url_atual}")
            _screenshot(page, "login_ERRO")
            return False

    except Exception as e:
        log.error(f"  ❌ Erro no login: {e}")
        _screenshot(page, "login_ERRO")
        return False


def fazer_login_displayforce_retry(page, tentativas: int = 3) -> bool:
    """Tenta login novamente em falhas transitórias de sessão/autenticação."""
    for i in range(1, tentativas + 1):
        if fazer_login_displayforce(page):
            return True
        log.warning(f"  Login falhou na tentativa {i}/{tentativas}")
        time.sleep(3)
    return False


def exportar_relatorio_cliente(page, nome_cliente: str, meses_offset: int = 0, escala_alvo: str = "Mês") -> bool:
    """Fluxo exato DisplayForce Panvel:
    meses_offset: 0=mês atual, 1=mês anterior, 2=dois meses atrás, -1=próximo mês
    escala_alvo: "Dia", "Semana", "Mês", "Trimestre", "Ano"
    1. Abre plataforma do cliente
    2. Clica em "Insights & dados" no menu lateral
    3. Clica no card "Visitors Insights"
    4. Seleciona escala "Mês" no dropdown de data
    5. Clica em "ENVIAR RELATÓRIO"
    6. Modal: marca "Inserir email manualmente" → preenche e-mail → clica ENVIAR
    """
    log.info(f"  Exportando relatório para: {nome_cliente}")
    nome_lower = nome_cliente.lower().strip()

    # ── Passo 1: Navega para a plataforma do cliente ───────────────────────
    page.goto(DISPLAYFORCE_PLATFORMS_URL, timeout=20_000)
    _aguardar_sem_networkidle(page, 2000)

    html = page.content()
    hrefs = re.findall(r'href=["\']([^"\']+)["\']', html)

    link_cliente = None
    for href in hrefs:
        if nome_lower in href.lower():
            link_cliente = href
            break
    if not link_cliente:
        for href in hrefs:
            partes = re.split(r"[/\-_]", href.lower())
            if any(nome_lower[:5] in p for p in partes if len(p) > 3):
                link_cliente = href
                break

    if link_cliente:
        url_full = link_cliente if link_cliente.startswith("http") else f"https://id.displayforce.ai{link_cliente}"
        log.info(f"  href encontrado para '{nome_cliente}': {url_full}")
        page.goto(url_full, timeout=20_000)
        _aguardar_sem_networkidle(page, 3000)
    else:
        try:
            page.locator(f"a:has-text('{nome_cliente}'), a:has-text('{nome_lower}')").first.click(timeout=8_000)
            _aguardar_sem_networkidle(page, 3000)
        except Exception:
            log.error(f"  ❌ '{nome_cliente}' não encontrado — pulando")
            return False

    log.info(f"  URL após abrir cliente: {page.url}")
    url_cliente_base = re.sub(r"/n/#.*|/#.*", "", page.url).rstrip("/")
    if not url_cliente_base.startswith("http"):
        url_cliente_base = f"https://{nome_lower}.displayforce.ai"
    log.info(f"  Base do cliente: {url_cliente_base}")
    _screenshot(page, f"01_cliente_{nome_cliente}")

    # ── Passo 2: Navega direto para visitors do mês correto via URL ──────────
    # Calcula o timestamp do mês alvo (meio do mês, fuso UTC)
    hoje = date.today()
    ano_alvo  = hoje.year
    mes_alvo  = hoje.month - meses_offset
    while mes_alvo <= 0:
        mes_alvo += 12
        ano_alvo -= 1
    ts_mes_ms = int(datetime(ano_alvo, mes_alvo, 15, 12, 0, 0, tzinfo=timezone.utc).timestamp() * 1000)
    url_visitors = f"{url_cliente_base}/n/#/stats/visitors?scale=month&date={ts_mes_ms}"
    log.info(f"  Navegando direto para {ano_alvo}-{mes_alvo:02d} via URL: {url_visitors}")
    try:
        page.goto(url_visitors, timeout=15_000)
        _aguardar_sem_networkidle(page, 3000)
        log.info(f"  URL visitors: {page.url}")
    except Exception:
        # Fallback sem parâmetro de data
        url_visitors = f"{url_cliente_base}/n/#/stats/visitors"
        log.warning(f"  Falha — tentando sem parâmetro de data: {url_visitors}")
        try:
            page.goto(url_visitors, timeout=15_000)
            _aguardar_sem_networkidle(page, 3000)
        except Exception:
            pass

    # Verifica se já está na página de visitors
    page_txt = (page.locator("body").text_content() or "").lower()
    tem_visitors = "visitors" in page_txt or "visitantes" in page_txt or "audiência" in page_txt or "atenção" in page_txt

    if not tem_visitors:
        # Tenta clicar no menu lateral "Insights & dados"
        for sel in [
            "a:has-text('Insights & dados')", "a:has-text('Insights')",
            "[class*='sidebar'] a:has-text('Insight')",
            "[class*='menu'] a:has-text('Insight')",
            "nav a:has-text('Insight')",
        ]:
            try:
                el = page.locator(sel).first
                if el.count() > 0:
                    el.click(timeout=5000)
                    log.info(f"  ✔ Clicado 'Insights & dados': {sel}")
                    _aguardar_sem_networkidle(page, 2000)
                    break
            except Exception:
                continue

        _screenshot(page, f"02_insights_menu_{nome_cliente}")

        # ── Passo 3: Clica no card "Visitors Insights" ─────────────────────
        log.info("  Clicando no card 'Visitors Insights'...")
        for sel in [
            "text='Visitors Insights'", "a:has-text('Visitors Insights')",
            "[class*='card']:has-text('Visitors')", "div:has-text('Visitors Insights'):visible",
            "a:has-text('Insights de Visitantes')", "text='Insights de Visitantes'",
        ]:
            try:
                el = page.locator(sel).first
                if el.count() > 0:
                    el.click(timeout=5000)
                    log.info(f"  ✔ Card Visitors Insights clicado: {sel}")
                    _aguardar_sem_networkidle(page, 3000)
                    break
            except Exception:
                continue

    _screenshot(page, f"03_visitors_{nome_cliente}")
    log.info(f"  URL Visitors Insights: {page.url}")

    # ── Passo 4: Seleciona escala no dropdown de data ─────────────────────
    # escala_alvo: "Dia", "Semana", "Mês", "Trimestre", "Ano"
    log.info(f"  Selecionando escala '{escala_alvo}'...")
    escala_alvo_lower = escala_alvo.lower()

    # Mapeamento pt/en para detectar se já está na escala certa
    escalas_map = {
        "dia": ("dia", "day"),
        "semana": ("semana", "week"),
        "mês": ("mês", "month", "mes"),
        "mes": ("mês", "month", "mes"),
        "trimestre": ("trimestre", "quarter"),
        "ano": ("ano", "year"),
    }
    escala_opcoes = escalas_map.get(escala_alvo_lower, (escala_alvo_lower,))

    escala_selecionada = False
    # Abre o dropdown (ele pode mostrar o valor atual: Dia, Semana, Mês etc.)
    for sel_dropdown in [
        "button:has-text('Mês')", "button:has-text('Month')",
        "button:has-text('Semana')", "button:has-text('Week')",
        "button:has-text('Dia')", "button:has-text('Day')",
        "button:has-text('Trimestre')", "button:has-text('Ano')",
        "[class*='select']:visible", "[class*='dropdown']:visible",
    ]:
        try:
            el = page.locator(sel_dropdown).first
            if el.count() == 0 or not el.is_visible():
                continue
            txt_atual = (el.text_content() or "").strip().lower()
            # Já está na escala certa?
            if txt_atual in escala_opcoes:
                log.info(f"  ✔ Escala já está em '{escala_alvo}'")
                escala_selecionada = True
                break
            # Abre o dropdown
            el.click(timeout=4000)
            log.info(f"  ✔ Dropdown de escala aberto (era '{txt_atual}')")
            time.sleep(1)
            # Seleciona a opção desejada
            for op_txt in [escala_alvo, escala_alvo.capitalize()]:
                for op_sel in [
                    f"li:has-text('{op_txt}')", f"option:has-text('{op_txt}')",
                    f"a:has-text('{op_txt}')", f"div:has-text('{op_txt}'):visible",
                    f"span:has-text('{op_txt}'):visible",
                ]:
                    try:
                        op_el = page.locator(op_sel).first
                        if op_el.count() > 0:
                            op_el.click(timeout=3000)
                            log.info(f"  ✔ Escala '{escala_alvo}' selecionada via: {op_sel}")
                            escala_selecionada = True
                            time.sleep(1)
                            break
                    except Exception:
                        continue
                if escala_selecionada:
                    break
            if escala_selecionada:
                break
        except Exception:
            continue

    if not escala_selecionada:
        log.info(f"  ℹ️  Dropdown de escala não encontrado — continuando com período padrão")

    # ── Confirma que o período correto está visível ────────────────────────
    # A navegação já foi feita via URL com timestamp — só verifica e loga
    time.sleep(1)
    page_txt_periodo = (page.locator("body").text_content() or "")
    # Procura texto do mês esperado na página
    meses_pt = ["", "Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    mes_esperado = meses_pt[mes_alvo] if 1 <= mes_alvo <= 12 else ""
    if mes_esperado and mes_esperado in page_txt_periodo:
        log.info(f"  ✔ Período confirmado na página: {mes_esperado} {ano_alvo}")
    else:
        log.warning(f"  ⚠️  Não foi possível confirmar o mês '{mes_esperado}' na página — continuando mesmo assim")

    _screenshot(page, f"04_periodo_{nome_cliente}")

    # ── Clica em "ENVIAR RELATÓRIO" ───────────────────────────────────────
    botoes = page.locator("a:visible, button:visible")
    textos_botoes = [botoes.nth(i).text_content() or "" for i in range(min(botoes.count(), 30))]
    log.info(f"  📋 Botões/links na página ({len(textos_botoes)}): {textos_botoes[:15]}")

    botao_export = None
    for i in range(botoes.count()):
        txt = (botoes.nth(i).text_content() or "").strip()
        if re.search(r"enviar.?relat|send.?report|export", txt, re.IGNORECASE):
            botao_export = botoes.nth(i)
            break

    if not botao_export:
        log.error("  ❌ Botão 'ENVIAR RELATÓRIO' não encontrado")
        return False

    botao_export.click()
    log.info("  ✔ Clicado: 'ENVIAR RELATÓRIO' (botão exportar)")
    time.sleep(2)

    # ── Modal: preenche e-mail e confirma ─────────────────────────────────
    _screenshot(page, f"10_modal_{nome_cliente}")

    # Loga estado do modal
    botoes_modal = page.locator("button:visible, a:visible")
    inputs_modal = page.locator("input:visible")
    textos_modal = [botoes_modal.nth(i).text_content() or "" for i in range(min(botoes_modal.count(), 20))]
    inputs_tipos = [inputs_modal.nth(i).get_attribute("type") or "text" for i in range(min(inputs_modal.count(), 10))]
    log.info(f"  📋 Modal aberto — botões: {textos_modal}")
    log.info(f"  📋 Modal aberto — inputs visíveis: {inputs_tipos}")

    # Clica em "Inserir email manualmente" (checkbox ou label)
    padroes_manual = [
        "label:has-text('manualmente')", "label:has-text('manually')",
        "label:has-text('Inserir')",
        "span:has-text('manualmente')", "div:has-text('manualmente'):visible",
        "button:has-text('manualmente')",
    ]
    checkbox_clicado = False
    for sel in padroes_manual:
        try:
            el = page.locator(sel).first
            if el.count() > 0 or page.locator(sel).count() > 0:
                page.locator(sel).first.click(timeout=4000)
                log.info(f"  ✔ Clicado elemento 'Inserir manualmente' com padrão: '{sel}'")
                checkbox_clicado = True
                time.sleep(0.8)
                break
        except Exception:
            continue

    if not checkbox_clicado:
        # Estratégia 2: checkbox puro
        try:
            cb = page.locator("input[type='checkbox']:visible").first
            if not cb.is_checked():
                cb.click()
                log.info("  ✔ Checkbox clicado (estratégia 2)")
                time.sleep(0.8)
        except Exception:
            log.info("  ℹ️  Checkbox 'Inserir email manualmente' não encontrado — tentando encontrar input direto")

    _screenshot(page, f"10b_apos_inserir_manual_{nome_cliente}")

    # Preenche o e-mail
    def _preencher_input_react(seletor: str, valor: str) -> bool:
        """Preenche um input React usando locator.evaluate — elemento passado automaticamente."""
        try:
            el = page.locator(seletor).first
            el.wait_for(state="visible", timeout=5000)
            el.triple_click()
            el.fill(valor)
            # locator.evaluate passa o elemento como 1º arg do JS automaticamente
            el.evaluate(
                """(el, val) => {
                    const setter = Object.getOwnPropertyDescriptor(
                        window.HTMLInputElement.prototype, 'value').set;
                    setter.call(el, val);
                    el.dispatchEvent(new Event('input', { bubbles: true }));
                    el.dispatchEvent(new Event('change', { bubbles: true }));
                }""",
                valor
            )
            return True
        except Exception:
            return False

    email_preenchido = False

    # Estratégia A: input[type=email] ou input[type=text] visível
    for sel in ("input[type='email']:visible", "input[type='text']:visible", "input:visible:not([type='checkbox'])"):
        try:
            inputs_vis = page.locator(sel)
            if inputs_vis.count() > 0:
                # Pega o input que está vazio ou tem e-mail padrão
                for i in range(inputs_vis.count()):
                    inp = inputs_vis.nth(i)
                    val_atual = inp.input_value()
                    if "@" not in val_atual or "globaltera" not in val_atual:
                        inp.triple_click()
                        inp.fill(RELATORIO_EMAIL)
                        # locator.evaluate — elemento passado automaticamente como 1º arg JS
                        inp.evaluate(
                            """(el, val) => {
                                const setter = Object.getOwnPropertyDescriptor(
                                    window.HTMLInputElement.prototype,'value').set;
                                setter.call(el, val);
                                el.dispatchEvent(new Event('input', { bubbles: true }));
                                el.dispatchEvent(new Event('change', { bubbles: true }));
                            }""",
                            RELATORIO_EMAIL
                        )
                        val_novo = inp.input_value()
                        log.info(f"  ✔ E-mail preenchido via '{sel}' — valor: '{val_novo}'")
                        email_preenchido = True
                        break
                if email_preenchido:
                    break
        except Exception:
            continue

    if not email_preenchido:
        # Fallback geral: percorre todos inputs visíveis
        log.warning("  ⚠️  Tentando fallback: percorrendo todos inputs visíveis")
        try:
            todos = page.locator(
                "input:visible:not([type='checkbox']):not([type='radio']):not([type='hidden'])"
                ":not([type='file']):not([type='submit']):not([type='button'])"
                ":not([type='number'])"
            )
            for i in range(todos.count()):
                inp = todos.nth(i)
                try:
                    inp.triple_click()
                    inp.fill(RELATORIO_EMAIL)
                    val = inp.input_value()
                    if RELATORIO_EMAIL[:5] in val:
                        log.info(f"  ✔ E-mail preenchido via fallback geral — valor: '{val}'")
                        email_preenchido = True
                        break
                except Exception:
                    continue
        except Exception:
            pass

    if not email_preenchido:
        # Último recurso: último input vazio
        log.warning("  ⚠️  Último recurso: preenchendo último input vazio disponível")
        try:
            ultimos = page.locator("input[type='text']:visible, input:not([type]):visible")
            for i in range(ultimos.count() - 1, -1, -1):
                try:
                    inp = ultimos.nth(i)
                    if not inp.input_value():
                        inp.fill(RELATORIO_EMAIL)
                        val = inp.input_value()
                        log.info(f"  ✔ E-mail preenchido no último input vazio — valor: '{val}'")
                        email_preenchido = True
                        break
                except Exception:
                    continue
        except Exception:
            pass

    if not email_preenchido:
        log.error("  ❌ CRÍTICO: Não foi possível preencher o e-mail no modal!")
        _screenshot(page, f"ERRO_email_{nome_cliente}")
        return False

    _screenshot(page, f"11_pre_confirmar_{nome_cliente}")

    # Clica em ENVIAR — escopo restrito ao modal-portal para não pegar botões atrás do overlay
    log.info("  Procurando botão ENVIAR dentro do modal...")
    clicou_confirmar = False

    # Estratégia 1: botão dentro do modal-portal (div que renderiza o modal acima de tudo)
    seletores_modal = [
        "#modal-portal button:has-text('ENVIAR')",
        "#modal-portal button:has-text('Enviar')",
        "#modal-portal button:has-text('SEND')",
        ".modal-portal button:has-text('ENVIAR')",
        ".modal-portal button:has-text('Enviar')",
        "[class*='modal'] button:has-text('ENVIAR')",
        "[class*='modal'] button:has-text('Enviar')",
        "[class*='dialog'] button:has-text('ENVIAR')",
    ]
    for sel in seletores_modal:
        try:
            el = page.locator(sel).first
            if el.count() > 0:
                el.click(timeout=5000)
                log.info(f"  ✔ Botão ENVIAR do modal clicado: {sel}")
                clicou_confirmar = True
                break
        except Exception:
            continue

    # Estratégia 2: JS direto no botão do modal (bypassa overlay)
    if not clicou_confirmar:
        try:
            clicou = page.evaluate("""() => {
                const portal = document.getElementById('modal-portal') ||
                               document.querySelector('.modal-portal') ||
                               document.querySelector('[class*="modal"]');
                if (!portal) return false;
                const btns = Array.from(portal.querySelectorAll('button'));
                const btn = btns.find(b => /enviar|send|confirm/i.test(b.textContent));
                if (btn) { btn.click(); return true; }
                return false;
            }""")
            if clicou:
                log.info("  ✔ Botão ENVIAR clicado via JS no modal-portal")
                clicou_confirmar = True
        except Exception as e:
            log.warning(f"  JS click falhou: {e}")

    # Estratégia 3: force=True no botão que o Playwright já encontrou
    if not clicou_confirmar:
        try:
            page.locator("button:has-text('ENVIAR'), button:has-text('Enviar')").last.click(
                timeout=5000, force=True
            )
            log.info("  ✔ Botão ENVIAR clicado com force=True")
            clicou_confirmar = True
        except Exception:
            pass

    # Estratégia 4: Enter como último recurso
    if not clicou_confirmar:
        page.keyboard.press("Enter")
        log.info("  ✔ Enter pressionado como confirmação final")

    time.sleep(2)
    _screenshot(page, f"12_pos_envio_{nome_cliente}")
    log.info(f"  ✅ Relatório enviado para {RELATORIO_EMAIL}")
    return True


# ── Lógica principal ──────────────────────────────────────────────────────────

def obter_periodo_atual() -> tuple[str, str]:
    """Retorna (inicio, fim) do mês atual em ISO format."""
    agora = datetime.now(timezone.utc)
    inicio = agora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    return inicio.isoformat(), agora.isoformat()


def executar_bot(headless: bool = True, filtro_clientes: list = None, meses_offset: int = 0, escala_alvo: str = "Mês"):
    """Executa um ciclo completo do bot."""
    if not _bot_lock.acquire(blocking=False):
        log.info("⏭️  Bot já em execução, aguardando próxima rodada...")
        return

    try:
        log.info("\n" + "=" * 60)
        log.info(f"  BOT DISPLAYFORCE — {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        log.info("=" * 60)

        # Tenta reenviar dados pendentes do ciclo anterior
        _reenviar_pendentes()

        # Busca clientes
        clientes = buscar_clientes()
        if not clientes:
            log.error("  Nenhum cliente ativo encontrado. Configure CLIENTES_FALLBACK no .env")
            return

        # Aplica filtro de clientes se especificado
        if filtro_clientes:
            filtro_lower = [f.lower().strip() for f in filtro_clientes]
            clientes_filtrados = [c for c in clientes if c["name"].lower().strip() in filtro_lower]
            if clientes_filtrados:
                log.info(f"  Filtro aplicado: processando apenas {[c['name'] for c in clientes_filtrados]}")
                clientes = clientes_filtrados
            else:
                log.warning(f"  Filtro '{filtro_clientes}' não bateu com nenhum cliente. Processando todos.")

        from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=headless)
            context = browser.new_context(
                viewport={"width": 1280, "height": 720},
                accept_downloads=True,
            )
            page = context.new_page()

            # Login único para todos os clientes
            ok_login = fazer_login_displayforce_retry(page, tentativas=3)
            if not ok_login:
                log.error("  Falha no login — abortando rodada")
                browser.close()
                return

            # Obtém o último UID do IMAP ANTES de solicitar relatórios
            # (para ignorar e-mails antigos)
            uid_inicio = 0
            try:
                imap = imaplib.IMAP4_SSL(IMAP_SERVER, IMAP_PORT)
                imap.login(IMAP_EMAIL, IMAP_PASSWORD)
                uid_inicio = _obter_ultimo_uid(imap)
                imap.logout()
            except Exception as e:
                log.warning(f"  Não foi possível obter UID inicial: {e}")

            # Para cada cliente: exporta relatório e processa
            for cliente in clientes:
                client_id   = str(cliente["id"])
                client_name = str(cliente["name"])
                log.info(f"\n  ── Processando: {client_name} ──")

                # Exporta relatório
                ok_export = exportar_relatorio_cliente(page, client_name, meses_offset=meses_offset, escala_alvo=escala_alvo)
                if not ok_export:
                    continue

                # Aguarda e-mail
                caminhos = baixar_relatorio_email(
                    uid_inicio=uid_inicio,
                    nome_cliente=client_name,
                    timeout_seg=TIMEOUT_EMAIL_SEG,
                )
                if not caminhos:
                    log.warning(f"  ⚠️  Relatório não recebido para '{client_name}'")
                    continue

                # Processa arquivos baixados
                registros: list[dict] = []
                for caminho in caminhos:
                    nome_arq = Path(caminho).name.lower()
                    try:
                        if "views" in nome_arq and caminho.endswith(".csv"):
                            registros.extend(processar_views_csv(caminho, client_id))
                        elif caminho.endswith(".csv"):
                            registros.extend(processar_views_csv(caminho, client_id))
                        elif caminho.endswith((".xlsx", ".xls")):
                            registros.extend(processar_excel(caminho, client_id))
                        else:
                            log.info(f"  Arquivo ignorado na extração de engajamento: {nome_arq}")
                    except Exception as e:
                        log.error(f"  Sem dados extraídos de: {caminho} — {e}")

                if registros:
                    upsert_campanhas(client_id, client_name, registros)
                else:
                    log.warning(f"  ⚠️  Nenhuma campanha salva para '{client_name}'")

                # Avança UID para não pegar o mesmo e-mail no próximo cliente
                try:
                    imap2 = imaplib.IMAP4_SSL(IMAP_SERVER, IMAP_PORT)
                    imap2.login(IMAP_EMAIL, IMAP_PASSWORD)
                    uid_inicio = _obter_ultimo_uid(imap2)
                    imap2.logout()
                except Exception:
                    pass

            browser.close()

        log.info("\n✅ Rodada concluída!\n")

    except Exception as e:
        import traceback
        log.error(f"  ❌ Erro inesperado na rodada: {e}")
        log.error(traceback.format_exc())
    finally:
        _bot_lock.release()


# ── Agendamento ───────────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Bot Campanhas DisplayForce")
    parser.add_argument("--agora", action="store_true",
                        help="Executa imediatamente em vez de aguardar o horário")
    parser.add_argument("--headless", type=str, default=None,
                        help="true/false — abre ou não o navegador visível (padrão: true)")
    parser.add_argument("--clientes", type=str, default=None,
                        help="Filtra clientes específicos separados por vírgula. Ex: --clientes Panvel,Assai")
    parser.add_argument("--mes-anterior", type=int, default=0,
                        help="Quantos meses atrás puxar (0=atual, 1=mês passado, 2=dois meses atrás)")
    parser.add_argument("--escala", type=str, default="Mês",
                        help="Escala do relatório: Dia, Semana, Mês, Trimestre, Ano (padrão: Mês)")
    args = parser.parse_args()

    headless = HEADLESS_DEFAULT
    if args.headless is not None:
        headless = args.headless.lower() not in ("false", "0", "no")

    # Filtro de clientes via linha de comando
    if args.clientes:
        filtro = [c.strip() for c in args.clientes.split(",") if c.strip()]
        log.info(f"  Filtro de clientes ativo: {filtro}")
    else:
        filtro = None

    meses_offset = getattr(args, "mes_anterior", 0) or 0
    if meses_offset:
        log.info(f"  Offset de meses: -{meses_offset} (clicará {meses_offset}x na seta <)")
    escala_alvo = getattr(args, "escala", "Mês") or "Mês"
    log.info(f"  Escala do relatório: {escala_alvo}")

    # Carrega configuração do Supabase (pode sobrescrever HORARIO_EXECUCAO)
    carregar_config_supabase()

    if args.agora:
        log.info("Modo --agora: executando imediatamente")
        executar_bot(headless=headless, filtro_clientes=filtro, meses_offset=meses_offset, escala_alvo=escala_alvo)
        return

    log.info("=" * 60)
    log.info(f"  Bot iniciado — execução diária às {HORARIO_EXECUCAO} (Brasília)")
    log.info("  Pressione Ctrl+C para encerrar")
    log.info("=" * 60)

    schedule.every().day.at(HORARIO_EXECUCAO).do(executar_bot, headless=headless)

    proximo = schedule.next_run()
    if proximo:
        falta = proximo - datetime.now()
        h, rem = divmod(int(falta.total_seconds()), 3600)
        m = rem // 60
        log.info(f"⏰ Agendamento ativo — próxima execução hoje/amanhã às {HORARIO_EXECUCAO} (falta {h}h {m}min)")

    while True:
        schedule.run_pending()
        time.sleep(30)


if __name__ == "__main__":
    main()
